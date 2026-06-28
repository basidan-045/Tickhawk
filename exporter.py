# =============================================================================
# exporter.py — Writes tick data to Excel using openpyxl.
#
# WHY openpyxl directly instead of pandas.to_excel()? More control over
# column widths, header formatting, and number formats without extra deps.
# For 200k rows it's also faster than going through pandas Excel writer.
# =============================================================================

import logging
import os
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import config
from data_store import TickStore

logger = logging.getLogger(__name__)


def _build_output_path(symbol: str, date_str: str) -> Path:
    """
    Build the full output file path for the Excel file.

    Args:
        symbol:   Stock symbol, e.g. "RELIANCE".
        date_str: Date string in DD-MM-YYYY format.

    Returns:
        Path object for the .xlsx output file.
    """
    output_dir = Path(config.OUTPUT_DIR)
    output_dir.mkdir(parents=True, exist_ok=True)
    filename = f"{symbol.upper()}_{date_str}.xlsx"
    return output_dir / filename


def export_to_excel(store: TickStore, symbol: str) -> Path:
    """
    Export all ticks from the store to an Excel file.

    Creates the output directory if it doesn't exist. Overwrites any
    previous backup written today (each auto-save refreshes the same file).

    Args:
        store:  TickStore instance to read from.
        symbol: Stock symbol for naming the file.

    Returns:
        Path to the written .xlsx file.
    """
    ticks = store.get_all_ticks()
    date_str = datetime.now().strftime("%d-%m-%Y")
    output_path = _build_output_path(symbol, date_str)

    if not ticks:
        logger.warning("No ticks to export yet — skipping Excel write.")
        return output_path

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{symbol.upper()} Ticks"

    # ── Header Row ──────────────────────────────────────────────────────────
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="1F4E79")  # Dark blue
    center_align = Alignment(horizontal="center")

    headers = ["Time", "Price (₹)"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # ── Data Rows ────────────────────────────────────────────────────────────
    for row_idx, tick in enumerate(ticks, start=2):
        ws.cell(row=row_idx, column=1, value=tick["time"])
        price_cell = ws.cell(row=row_idx, column=2, value=tick["price"])
        price_cell.number_format = '#,##0.00'

    # ── Column Widths ────────────────────────────────────────────────────────
    ws.column_dimensions[get_column_letter(1)].width = 18  # Time column
    ws.column_dimensions[get_column_letter(2)].width = 15  # Price column

    # ── Freeze top row so headers stay visible while scrolling ───────────────
    ws.freeze_panes = "A2"

    wb.save(output_path)
    logger.info(f"Excel saved: {output_path} ({len(ticks):,} rows)")
    print(f"\n[SAVED] Excel → {output_path} ({len(ticks):,} ticks)")

    return output_path
